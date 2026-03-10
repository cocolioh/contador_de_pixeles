from PIL import Image
from openpyxl import Workbook
import openpyxl, os
from openpyxl.styles import PatternFill

path = r"PATH HERE"

if path == "PATH HERE":
    print("Por favor escribir la ruta de la imagen directo en el archivo .py")
    print("Please write the image path directly onto the .py file")
    exit()


# http://www.psychocodes.in/rgb-to-hex-conversion-and-hex-to-rgb-conversion-in-python.html
def rgb2hex(r,g,b):
    hex = "#{:02x}{:02x}{:02x}".format(r,g,b)
    return hex

'''class Color:
    color = "color"
    alpha = 0
    cantidad = 0

    def __init__(self, col, a):
        self.color = col
        self.alpha = 0
        self.cant = 0
'''



im = Image.open(path)

print(im.format, "ancho:", im.size[0], "alto:", im.size[1], im.mode)

ancho = im.size[0]
alto = im.size[1]

# Getpixel obtiene formato (red, green, blue, alpha)
# Devuelve una lista en ese orden

# color = im.getpixel([0,0])

# print(color[1])


colores = []
alphas = []
cantidades = []

for i in range(ancho):
    for j in range(alto):
        print("Analizando pixel ", i, ", ", j)
        temp = im.getpixel([i,j])
        print("Color pixel actual:", temp)
        color_hex = (rgb2hex(temp[0],temp[1],temp[2])).upper() # Sera #FFFFFF
        

        if len(temp) == 4:
            print("Alpha en rgb:", temp[3])
            # x = int(temp[3]/255)*100
            # print("Alpha: ", x)
            hex_alpha = (hex(temp[3])[2:]).upper() # sera FF o lo que sea
        else:
            hex_alpha = "FF"
        #color_alpha = temp[3]

        color_hex = (color_hex.split("#"))[1]

        if len(hex_alpha) == 1:
            color_hex = "0" + hex_alpha + color_hex
        else:
            color_hex = hex_alpha + color_hex


        print("Color pixel actual en hex:", color_hex)

        if color_hex not in colores:
            colores.append(color_hex)
            #alphas.append(color_alpha)
            cantidades.append(1)
        else:
            indice = colores.index(color_hex)
            cantidades[indice] += 1

# print(colores)
# print(cantidades)

wb = Workbook()

ws = wb.active

ws['A1'] = 'Color'
#ws['B1'] = 'Alpha'
ws['B1'] = 'Cantidad'


for color in colores:
    indice = colores.index(color)

    cell = ws.cell(indice+2, 1, color)
    fill = PatternFill(fgColor=color, fill_type='solid')
    # fill = PatternFill(fgColor="#FFFFFF", fill_type='solid')
    cell.fill = fill
    
    #cell = ws.cell(indice+2, 2, alphas[indice])
    
    cell = ws.cell(indice+2, 2, cantidades[indice])

separador = "\\"

path_separado = path.split(separador)

nombre_archivo = path_separado[-1]

nombre_archivo_separado = nombre_archivo.split(".")

nombre_sheet = nombre_archivo_separado[0] + ".xlsx"

wb.save(nombre_sheet)

print("Todos los colores obtenidos exitosamente.")


        
            
        
        
        




